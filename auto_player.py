#!/usr/bin/env python3
"""
Auto-Player with Multi-Account Rotation, Auto-Like, Auto-Comment & Auto-Subscribe
================================================================================
Runs browser automation on YouTube to:
  1. Discover multiple YouTube brand/channel accounts within a single Chrome profile.
  2. Rotate through accounts one by one every 15 to 25 minutes (configurable).
  3. Browse and watch trending videos (global trending feeds & niche trending, videos/shorts).
  4. Automatically like videos with human-like timing and probability.
  5. Automatically leave natural, customizable comments from a comment pool.
  6. Automatically subscribe to creators with randomized probability.
  7. Log all watched videos, likes, comments, subscriptions, and metrics to Excel & CSV.

Can be run directly via CLI or controlled via the Bot UI (`python bot_ui.py` / `bot_UI.bat`).
"""
from __future__ import annotations

import json
import os
import random
import re
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

HERE = Path(__file__).resolve().parent
CONFIG = HERE / "config.json"
STOP_FILE = HERE / "stop.txt"

DEFAULT_COMMENTS = [
    "Great video, thanks for sharing!",
    "Really well explained, enjoyed this one 👍",
    "Super helpful content, keep it up!",
    "Loved the breakdown on this. Very insightful!",
    "Quality video right here 🔥",
    "This was really informative, thanks!",
    "Awesome presentation and pacing 👌",
    "Nice video! Learned something new today.",
    "Great editing and clear explanations!",
    "Solid content as always 👏",
    "Subscribed! Looking forward to more videos.",
    "The explanation at the beginning was super clear 🙌",
]

DEFAULTS = {
    "niche": "",
    "use_real_chrome": False,
    "connect_cdp": False,
    "cdp_url": "http://localhost:9222",
    "chrome_profile_dir": "Default",
    "output_dir": "output",
    "play_seconds_per_video": 40,
    "play_jitter": 12,
    "break_every": 8,
    "break_seconds": [90, 180],
    "max_videos": 0,
    "account_rotation": {
        "enabled": True,
        "rotate_minutes": [15, 25],
        "accounts": [],
    },
    "trending": {
        "enabled": True,
        "content_type": "both",  # "both", "videos", "shorts"
        "source": "trending_and_niche",
        "categories": ["trending", "gaming", "music"],
    },
    "engagement": {
        "auto_like": True,
        "like_probability": 0.8,
        "auto_comment": True,
        "comment_probability": 0.5,
        "auto_subscribe": True,
        "subscribe_probability": 0.25,
        "comment_pool": DEFAULT_COMMENTS,
    },
}


def load_cfg() -> dict:
    cfg = dict(DEFAULTS)
    try:
        if CONFIG.exists():
            user_cfg = json.loads(CONFIG.read_text(encoding="utf-8"))
            for k, v in user_cfg.items():
                if isinstance(v, dict) and isinstance(cfg.get(k), dict):
                    cfg[k].update(v)
                else:
                    cfg[k] = v
    except Exception as e:
        print(f"[Warning] Failed to read {CONFIG.name}: {e}")
    return cfg


def real_chrome_profile_dir():
    home = Path.home()
    candidates = [
        home / "AppData" / "Local" / "Google" / "Chrome" / "User Data",
        home / "Library" / "Application Support" / "Google" / "Chrome",
        home / ".config" / "google-chrome",
        home / ".config" / "chromium",
    ]
    for c in candidates:
        if c.is_dir():
            return str(c)
    return None


def _ignore_copy_errors(src, names):
    """Copy everything except transient lock files & logs (these change constantly)."""
    skip = {"SingletonLock", "SingletonCookie", "SingletonSocket", "LOG", "LOG.old",
            "lockfile", "LOCK", "Local State", "GPU Cache", "GPUCache", "Cached Data",
            "Code Cache", "DawnCache", "GrShaderCache", "ShaderCache", "History-journal"}
    return [n for n in names if n in skip]


def clone_real_profile(real_user_data_dir: str, profile_name: str) -> Path:
    """
    Make a writable, bot-owned COPY of a Chrome profile so Playwright can launch it
    WITHOUT fighting the live profile's singleton lock / Chrome 136 restrictions.

    Chrome refuses to let Playwright reliably drive a profile that the real Chrome
    process owns or has locked (hence 'Target page, context or browser has been closed'
    / launch timeouts). Copying the profile's cookies, local storage and settings into a
    fresh user-data dir preserves ALL signed-in accounts so the bot can rotate them,
    while leaving the user's real Chrome untouched.

    Returns the path to the cloned user-data dir (whose 'Default' profile holds the
    copied login). Raises RuntimeError on failure.
    """
    src_root = Path(real_user_data_dir)
    src_profile = src_root / profile_name
    if not src_profile.is_dir():
        raise RuntimeError("Profile '%s' not found under %s" % (profile_name, src_root))

    dst_root = HERE / "cloned_profile"
    # Fresh copy each run so we don't re-use a possibly corrupt/partial snapshot.
    if dst_root.exists():
        import shutil as _sh
        try:
            _sh.rmtree(dst_root, ignore_errors=True)
        except Exception:
            pass
    dst_root.mkdir(parents=True, exist_ok=True)

    import shutil
    # Copy 'Local State' (holds the cookie encryption key) and the profile folder into
    # the clone's Default slot. Chrome reads the Default profile when no
    # --profile-directory is given.
    try:
        local_state = src_root / "Local State"
        if local_state.is_file():
            shutil.copy2(local_state, dst_root / "Local State")
    except Exception:
        pass

    default_dir = dst_root / "Default"
    default_dir.mkdir(parents=True, exist_ok=True)
    shutil.copytree(
        src_profile,
        default_dir,
        dirs_exist_ok=True,
        ignore=_ignore_copy_errors,
    )
    # Chrome also stores some shared state at the User Data root; copy "Preferences"
    # for the account if present, but it's optional.
    return dst_root


def chrome_is_running() -> bool:
    """Cheap check: is a Chrome/Chromium browser process already running?"""
    try:
        if sys.platform == "win32":
            out = subprocess.run(["tasklist"], capture_output=True, text=True, timeout=10).stdout
            return "chrome.exe" in out.lower()
        out = subprocess.run(["pgrep", "-l", "chrome"], capture_output=True, text=True, timeout=10).stdout
        return "chrome" in out.lower()
    except Exception:
        return False


def _chrome_lock_roots():
    """Candidate profile root directories that may hold Chrome Singleton lock files."""
    home = Path.home()
    return [
        home / "AppData" / "Local" / "Google" / "Chrome" / "User Data",
        home / "Library" / "Application Support" / "Google" / "Chrome",
        home / ".config" / "google-chrome",
        home / ".config" / "chromium",
    ]


def _remove_stale_locks():
    """Delete Chrome's Singleton lock files so a profile is not seen as 'in use'."""
    for root in _chrome_lock_roots():
        if not root.exists():
            continue
        candidates = [root] + ([p for p in root.iterdir() if p.is_dir()] if root.is_dir() else [])
        try:
            candidates = list(dict.fromkeys(candidates))
        except Exception:
            pass
        for folder in candidates:
            for lock in ("SingletonLock", "SingletonCookie", "SingletonSocket"):
                try:
                    (folder / lock).unlink(missing_ok=True)
                except Exception:
                    pass


def _chrome_procs():
    """Return True if any Chrome/Chromium process is running."""
    try:
        if sys.platform == "win32":
            out = subprocess.run(["tasklist"], capture_output=True, text=True, timeout=10).stdout
            return "chrome.exe" in out.lower()
        out = subprocess.run(["pgrep", "-l", "chrome"], capture_output=True, text=True, timeout=10).stdout
        return "chrome" in out.lower()
    except Exception:
        return False


def close_chrome():
    """Force-close every Chrome/Chromium process and verify it is actually gone.

    This is the key to fixing the 'launch_persistent_context: Timeout' error. Chrome is
    a single-instance browser per user-data-dir: if ANY chrome.exe is still alive (even a
    background/helper process, or a leftover instance from a previous bot run that hung),
    a new launch simply hands off to the existing one and Playwright never receives a
    connection - so it times out after 60s. We therefore kill ALL chrome.exe processes,
    wait until NONE remain, and delete the stale Singleton lock files the kill leaves behind.
    """
    try:
        for attempt in range(6):
            if not _chrome_procs():
                break
            try:
                if sys.platform == "win32":
                    # /T kills child processes, /F forces, /IM matches the image name.
                    subprocess.run(["taskkill", "/F", "/T", "/IM", "chrome.exe"],
                                   capture_output=True, timeout=20)
                    subprocess.run(["taskkill", "/F", "/T", "/IM", "chrome_proxy.exe"],
                                   capture_output=True, timeout=10)
                    subprocess.run(["taskkill", "/F", "/T", "/IM", "msedge.exe"],
                                   capture_output=True, timeout=5)
                else:
                    subprocess.run(["pkill", "-9", "-f", "chrome"], capture_output=True, timeout=20)
            except Exception:
                pass
            time.sleep(2)

        # Delete stale Singleton lock files left behind by the forced kill.
        _remove_stale_locks()
        time.sleep(1)
    except Exception:
        pass


def stopped() -> bool:
    return STOP_FILE.exists()


def _is_signed_in(page) -> bool:
    """Best-effort check: is the session logged into a YouTube/Google account?"""
    try:
        avatar = page.query_selector(
            "button#avatar-btn, ytd-topbar-menu-button-renderer #avatar-btn, yt-img-shadow#avatar"
        )
        if avatar and avatar.is_visible():
            return True
        # If we can still see a "Sign in" button we are NOT signed in.
        sign_in = page.query_selector(
            "a[href*='ServiceLogin'], ytd-button-renderer a[aria-label*='Sign in' i]"
        )
        if sign_in and sign_in.is_visible():
            return False
    except Exception:
        pass
    return False


def wait_for_login(page, status_callback=None, max_wait: int = 180) -> bool:
    """
    Blocks (up to max_wait seconds) until the user has signed into YouTube in the
    opened browser window. Returns True when signed in, False on timeout.

    Works in BOTH CLI and Web-UI mode: instead of relying on a console `input()`
    (which dead-locks when launched from the dashboard), we poll the page and emit
    status events so the UI can tell the user to log in.
    """
    def emit(event_type, **data):
        if status_callback:
            try:
                status_callback({"type": event_type, "timestamp": datetime.now().isoformat(), **data})
            except Exception:
                pass

    print("\nBrowser is open. If this is your first run, please log into your Google/YouTube")
    print("account in the browser window that just opened. The bot will continue automatically")
    print("once it detects that you are signed in.\n")

    if _is_signed_in(page):
        print("[Login] Account already signed in. Continuing...")
        return True

    emit("status", status="Waiting for login — please sign in to YouTube in the browser window...")
    print("[Login] Waiting for you to sign in... (up to %d seconds)" % max_wait)

    start = time.time()
    while time.time() - start < max_wait:
        if stopped():
            return False
        if _is_signed_in(page):
            print("[Login] Signed-in detected. Starting automation...")
            emit("status", status="Login detected — starting automation...")
            return True
        time.sleep(3)

    print("[Login] No login detected within %d seconds. Continuing as guest / single account." % max_wait)
    emit("status", status="No login detected within timeout — continuing as guest.")
    return False


def _chrome_exe():
    if sys.platform == "win32":
        cands = [
            Path(os.environ.get("ProgramFiles", r"C:\Program Files")) / "Google/Chrome/Application/chrome.exe",
            Path(os.environ.get("ProgramFiles(x86)", r"C:\Program Files (x86)")) / "Google/Chrome/Application/chrome.exe",
            Path.home() / "AppData/Local/Google/Chrome/Application/chrome.exe",
        ]
    else:
        cands = [
            Path("/usr/bin/google-chrome"),
            Path("/usr/bin/chromium-browser"),
            Path("/usr/bin/chromium"),
            Path("/Applications/Google Chrome.app/Contents/MacOS/Google Chrome"),
        ]
    for c in cands:
        if c.exists():
            return str(c)
    return "chrome"


def _ensure_browser() -> str:
    """
    Verify Playwright's bundled Chromium is installed. Returns the browser path or
    raises a RuntimeError with a clear message if it is missing. This lets the bot
    fail fast with instructions instead of hanging silently when the window doesn't open.
    """
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            try:
                return p.chromium.executable_path
            except Exception as e:
                raise RuntimeError(
                    "Playwright Chromium is not installed. Run the launcher again (it installs "
                    "it) or run manually:  python -m playwright install chromium\nDetails: %s" % str(e)[:120]
                )
    except RuntimeError as e:
        raise
    except Exception as e:
        raise RuntimeError(
            "Playwright is not available. Install it with:  pip install -r requirements.txt\n"
            "then:  python -m playwright install chromium\nDetails: %s" % str(e)[:120]
        )


def _stealth_args(extra=None):
    """
    Chrome command-line args that stop Google's "This browser or app may not be secure"
    sign-in block. Playwright's default Chromium is detected as automated, so we launch
    the user's REAL Google Chrome (channel='chrome') with --disable-blink-features=
    AutomationControlled and without --enable-automation, so login works.
    """
    args = [
        "--disable-blink-features=AutomationControlled",
        "--disable-infobars",
        "--no-first-run",
        "--no-default-browser-check",
        "--start-maximized",
        "--disable-features=TranslateUI",
    ]
    if extra:
        args.extend(extra)
    return args


def _stealth_init_script():
    """
    JavaScript run on every page to hide automation markers Google checks for.
    """
    return """
    Object.defineProperty(navigator, 'webdriver', { get: () => undefined });
    window.chrome = window.chrome || { runtime: {} };
    Object.defineProperty(navigator, 'languages', { get: () => ['en-US', 'en'] });
    Object.defineProperty(navigator, 'plugins', { get: () => [1, 2, 3, 4, 5] });
    const origQuery = window.navigator.permissions.query.bind(window.navigator.permissions);
    window.navigator.permissions.query = (params) => (
        params.name === 'notifications'
            ? Promise.resolve({ state: Notification.permission })
            : origQuery(params)
    );
    """


def launch_persistent_browser(p, user_data_dir, headless=False, extra_args=None, channel="chrome"):
    """
    Launch a persistent Chrome context that can pass Google's sign-in security check.
    Uses the REAL installed Chrome (channel='chrome') with stealth flags, so the user
    can log into their Google account without "This browser or app may not be secure".
    Returns a BrowserContext.
    """
    ctx = p.chromium.launch_persistent_context(
        str(user_data_dir),
        channel=channel,
        headless=headless,
        args=_stealth_args(extra_args),
        ignore_default_args=["--enable-automation"],
        viewport=None,
        locale="en-US",
        timeout=90000,
    )
    try:
        ctx.add_init_script(_stealth_init_script())
    except Exception:
        pass
    return ctx


def _dismiss_banners(page):
    """Dismisses consent, cookie, or info overlays."""
    selectors = [
        "button[aria-label*='Accept all']",
        "button[aria-label*='Accept the use']",
        "tp-yt-paper-button#agree-button",
        "ytd-button-renderer button#agree-button",
        "button#accept-button",
        "yt-button-shape button[aria-label*='Accept']",
        "button[aria-label*='Reject all']",
        "ytd-button-renderer:has-text('No thanks') button",
        "button[aria-label='Dismiss']",
        "tp-yt-paper-dialog button[aria-label*='Dismiss' i]",
    ]
    for sel in selectors:
        try:
            b = page.query_selector(sel)
            if b and b.is_visible():
                b.click()
                time.sleep(1)
                return
        except Exception:
            pass


def _scrape_meta(page) -> dict:
    """Best-effort: read view count, channel name, and duration off the active page."""
    meta = {"views": "", "channel": "", "duration": ""}
    # 1. View count
    for sel in (
        "ytd-watch-metadata #count yt-formatted-string",
        "ytd-watch-metadata #count",
        "ytd-video-view-count",
        "ytd-video-primary-info-renderer #view-count",
        "yt-content-view-engine-view-model #count",
        "span.yt-view-count-renderer",
    ):
        try:
            el = page.query_selector(sel)
            if el:
                txt = (el.text_content() or "")
                m = re.search(r"([\d.,]+)", txt.replace("views", "").replace("watching", ""))
                if m:
                    num = m.group(1).replace(",", "").replace(".", "")
                    if num.isdigit():
                        meta["views"] = int(num)
                        break
        except Exception:
            continue

    # 2. Channel name
    for sel in (
        "ytd-channel-name a#channel-name",
        "ytd-channel-name #channel-name",
        "yt-formatted-string#channel-name",
        "ytd-video-owner-renderer #channel-name a",
        "#owner #channel-name",
        "ytd-reel-video-renderer[is-active] ytd-channel-name a",
    ):
        try:
            el = page.query_selector(sel)
            if el and (el.text_content() or "").strip():
                meta["channel"] = el.text_content().strip()
                break
        except Exception:
            continue

    # 3. Duration from the <video> element
    try:
        el = page.query_selector("video")
        if el:
            d = el.get_attribute("duration")
            if d:
                meta["duration"] = int(float(d))
    except Exception:
        pass
    return meta


def fmt_len(s) -> str:
    try:
        s = int(s)
    except (TypeError, ValueError):
        return ""
    m, sec = divmod(s, 60)
    h, m = divmod(m, 60)
    return ("%d:%02d:%02d" % (h, m, sec)) if h else ("%d:%02d" % (m, sec))


def _flag(views) -> str:
    try:
        v = int(views)
    except (TypeError, ValueError):
        return ""
    if v >= 250000:
        return "Trending — high reach"
    if v >= 25000:
        return "Growing — strong engagement"
    if v < 1000:
        return "Small creator — supportive reach"
    return "Standard"


# --------------------------------------------------------------------------
# Multi-Account Detection & Switching
# --------------------------------------------------------------------------
def discover_available_accounts(page) -> list[dict]:
    """
    Scans the YouTube profile session to discover all channel/brand accounts.
    Returns list of dicts: [{'index': 0, 'name': 'Channel A', 'handle': '@chA'}, ...]
    """
    print("\n[Accounts] Detecting linked YouTube channels in profile...")

    # Strategy 1: Navigate to channel switcher URL
    try:
        page.goto("https://www.youtube.com/channel_switcher", timeout=40000)
        page.wait_for_load_state("domcontentloaded")
        time.sleep(3)
        _dismiss_banners(page)

        items = page.query_selector_all("ytd-account-item-renderer, tp-yt-paper-item")
        if items:
            raw = page.evaluate("""() => {
                const results = [];
                const els = document.querySelectorAll('ytd-account-item-renderer, tp-yt-paper-item');
                els.forEach((el, idx) => {
                    const title = el.querySelector('#channel-title, yt-formatted-string#channel-title, #account-name, .channel-title');
                    const handle = el.querySelector('#email, yt-formatted-string#email, .email');
                    const name = title ? title.textContent.trim() : '';
                    if (name && !name.toLowerCase().includes('create a channel') && !name.toLowerCase().includes('add account')) {
                        results.push({
                            index: idx,
                            name: name,
                            handle: handle ? handle.textContent.trim() : ''
                        });
                    }
                });
                return results;
            }""")
            if raw and len(raw) > 0:
                print(f"[Accounts] Found {len(raw)} account(s) via channel_switcher:")
                for a in raw:
                    print(f"   [{a['index'] + 1}] {a['name']} {('(' + a['handle'] + ')') if a.get('handle') else ''}")
                return raw
    except Exception as ex:
        print(f"[Accounts] Notice during channel_switcher scan: {str(ex)[:80]}")

    # Strategy 2: Check Avatar Menu
    try:
        page.goto("https://www.youtube.com", timeout=40000)
        page.wait_for_load_state("domcontentloaded")
        time.sleep(2)
        _dismiss_banners(page)

        avatar = page.query_selector("button#avatar-btn, ytd-topbar-menu-button-renderer #avatar-btn, yt-img-shadow#avatar")
        if avatar:
            avatar.click()
            time.sleep(1.5)
            switch_btn = page.query_selector("ytd-compact-link-renderer:has-text('Switch account'), tp-yt-paper-item:has-text('Switch account')")
            if switch_btn:
                switch_btn.click()
                time.sleep(1.5)
                raw = page.evaluate("""() => {
                    const results = [];
                    const els = document.querySelectorAll('ytd-account-item-renderer');
                    els.forEach((el, idx) => {
                        const title = el.querySelector('#channel-title, yt-formatted-string#channel-title, #account-name');
                        const handle = el.querySelector('#email, yt-formatted-string#email');
                        const name = title ? title.textContent.trim() : '';
                        if (name) {
                            results.push({
                                index: idx,
                                name: name,
                                handle: handle ? handle.textContent.trim() : ''
                            });
                        }
                    });
                    return results;
                }""")
                if raw and len(raw) > 0:
                    print(f"[Accounts] Found {len(raw)} account(s) via Avatar Menu:")
                    for a in raw:
                        print(f"   [{a['index'] + 1}] {a['name']} {('(' + a['handle'] + ')') if a.get('handle') else ''}")
                    return raw
    except Exception as ex:
        print(f"[Accounts] Notice during avatar menu scan: {str(ex)[:80]}")

    print("[Accounts] Single default account / guest session active.")
    return [{"index": 0, "name": "Active Profile Account", "handle": ""}]


def switch_to_account(page, target_idx: int, target_name: str = "") -> str:
    """
    Switches the active YouTube session to the specified channel/account.
    Returns the resolved account name.
    """
    label = target_name or f"Account #{target_idx + 1}"
    print(f"\n" + "=" * 50)
    print(f"[Account Switcher] Switching to: {label}...")
    print("=" * 50)

    # Strategy 1: via channel_switcher page
    try:
        page.goto("https://www.youtube.com/channel_switcher", timeout=40000)
        page.wait_for_load_state("domcontentloaded")
        time.sleep(3)
        _dismiss_banners(page)

        switched = page.evaluate("""(target) => {
            const els = Array.from(document.querySelectorAll('ytd-account-item-renderer, tp-yt-paper-item'));
            if (!els || els.length === 0) return null;
            
            if (typeof target.name === 'string' && target.name.trim().length > 0) {
                const targetLower = target.name.toLowerCase();
                for (let i = 0; i < els.length; i++) {
                    const title = els[i].querySelector('#channel-title, yt-formatted-string#channel-title, #account-name');
                    const name = title ? title.textContent.trim() : '';
                    if (name.toLowerCase().includes(targetLower) || targetLower.includes(name.toLowerCase())) {
                        els[i].click();
                        return name;
                    }
                }
            }
            const idx = target.idx % els.length;
            const title = els[idx].querySelector('#channel-title, yt-formatted-string#channel-title, #account-name');
            const name = title ? title.textContent.trim() : `Account #${idx + 1}`;
            els[idx].click();
            return name;
        }""", {"idx": target_idx, "name": target_name})

        if switched:
            time.sleep(4)
            page.wait_for_load_state("domcontentloaded")
            print(f"[Account Switcher] Switched successfully to: '{switched}'")
            return switched
    except Exception as e:
        print(f"[Account Switcher] Notice via channel_switcher: {str(e)[:80]}")

    # Strategy 2: via Topbar Avatar Dropdown
    try:
        page.goto("https://www.youtube.com", timeout=40000)
        page.wait_for_load_state("domcontentloaded")
        time.sleep(2)
        _dismiss_banners(page)

        avatar = page.query_selector("button#avatar-btn, ytd-topbar-menu-button-renderer #avatar-btn, yt-img-shadow#avatar")
        if avatar:
            avatar.click()
            time.sleep(1.5)
            switch_btn = page.query_selector("ytd-compact-link-renderer:has-text('Switch account'), tp-yt-paper-item:has-text('Switch account')")
            if switch_btn:
                switch_btn.click()
                time.sleep(1.5)
                items = page.query_selector_all("ytd-account-item-renderer")
                if items:
                    sel_item = items[target_idx % len(items)]
                    sel_item.click()
                    time.sleep(4)
                    page.wait_for_load_state("domcontentloaded")
                    print(f"[Account Switcher] Switched to item #{target_idx + 1} via Avatar menu.")
                    return label
    except Exception as ex:
        print(f"[Account Switcher] Notice via avatar dropdown: {str(ex)[:80]}")

    return label


# --------------------------------------------------------------------------
# Automated Liking, Commenting & Subscribing
# --------------------------------------------------------------------------
def perform_like(page, probability: float = 0.8) -> bool:
    """Detects and clicks the Like button if not already liked."""
    if random.random() > probability:
        return False
    try:
        time.sleep(random.uniform(1.5, 3.0))
        like_selectors = [
            "segmented-like-dislike-button-view-model like-button-view-model button",
            "like-button-view-model button",
            "ytd-segmented-like-dislike-button-renderer button:first-child",
            "ytd-like-button-renderer button",
            "button[aria-label*='like this video' i]",
            "button[aria-label*='like' i][aria-label*='video' i]",
            "ytd-reel-video-renderer[is-active] #like-button button",
            "#like-button button",
        ]
        for sel in like_selectors:
            btn = page.query_selector(sel)
            if btn and btn.is_visible():
                pressed = btn.get_attribute("aria-pressed")
                if pressed and pressed.lower() == "true":
                    return True  # Already liked
                label = (btn.get_attribute("aria-label") or "").lower()
                if "dislike" in label:
                    continue
                btn.scroll_into_view_if_needed()
                btn.click()
                time.sleep(random.uniform(0.5, 1.2))
                return True
    except Exception:
        pass
    return False


def perform_comment(page, comment_pool: list[str], probability: float = 0.5) -> str:
    """Scrolls to the comment box, writes a natural comment, and submits it."""
    if not comment_pool or random.random() > probability:
        return ""
    try:
        # 1. Scroll down to trigger comments loading
        page.evaluate("""() => {
            const c = document.querySelector('#comments, ytd-comments');
            if (c) c.scrollIntoView({ behavior: 'smooth', block: 'center' });
            else window.scrollBy(0, 500);
        }""")
        time.sleep(random.uniform(2.0, 3.5))

        # Check for comments disabled notice
        disabled = page.query_selector("ytd-comments #message:has-text('Comments are turned off'), ytd-message-renderer")
        if disabled and disabled.is_visible():
            return ""

        # 2. Click placeholder to activate editor
        placeholder_selectors = [
            "#simplebox-placeholder",
            "#placeholder-area",
            "ytd-comment-simplebox-renderer #simplebox-placeholder",
            "#simplebox-container",
        ]
        for sel in placeholder_selectors:
            box = page.query_selector(sel)
            if box and box.is_visible():
                box.scroll_into_view_if_needed()
                box.click()
                break
        time.sleep(random.uniform(1.0, 2.0))

        # 3. Locate editable content area
        editor_selectors = [
            "div#contenteditable-root",
            "#comment-dialog div#contenteditable-root",
            "ytd-commentbox #contenteditable-root",
            "div[contenteditable='true']#contenteditable-root",
            "div#contenteditable-textarea",
        ]
        editor = None
        for sel in editor_selectors:
            el = page.query_selector(sel)
            if el and el.is_visible():
                editor = el
                break

        if not editor:
            return ""

        chosen_comment = random.choice(comment_pool)
        editor.click()
        time.sleep(0.4)
        # Type naturally with human-like key intervals
        page.keyboard.type(chosen_comment, delay=random.randint(35, 85))
        time.sleep(random.uniform(1.0, 2.0))

        # 4. Click Submit / Comment button
        submit_selectors = [
            "#submit-button button",
            "ytd-button-renderer#submit-button button",
            "button[aria-label*='Comment' i]",
            "ytd-commentbox #submit-button yt-button-shape button",
            "#comment-dialog #submit-button button",
        ]
        for sel in submit_selectors:
            sub = page.query_selector(sel)
            if sub and sub.is_visible():
                if sub.get_attribute("disabled") is not None:
                    time.sleep(1.0)
                sub.click()
                time.sleep(random.uniform(1.5, 2.5))
                return chosen_comment

    except Exception:
        pass
    return ""


def perform_subscribe(page, probability: float = 0.25) -> bool:
    """Detects and clicks the Subscribe button if not already subscribed."""
    if random.random() > probability:
        return False
    try:
        time.sleep(random.uniform(1.5, 3.0))
        sub_selectors = [
            "ytd-watch-metadata #subscribe-button button",
            "ytd-subscribe-button-renderer button",
            "#subscribe-button yt-button-shape button",
            "#subscribe-button button",
            "ytd-reel-video-renderer[is-active] #subscribe-button button",
            "button[aria-label*='Subscribe to' i]",
            "button[aria-label*='Subscribe' i]",
        ]
        for sel in sub_selectors:
            btn = page.query_selector(sel)
            if btn and btn.is_visible():
                btn_text = (btn.text_content() or "").strip().lower()
                aria_label = (btn.get_attribute("aria-label") or "").lower()

                # Check if already subscribed
                if (
                    "subscribed" in btn_text
                    or "unsubscribe" in btn_text
                    or "subscribed" in aria_label
                    or "unsubscribe" in aria_label
                ):
                    return False

                # Verify it is a subscribe action
                if "subscribe" in btn_text or "subscribe" in aria_label:
                    btn.scroll_into_view_if_needed()
                    btn.click()
                    time.sleep(random.uniform(0.8, 1.8))

                    # Dismiss any membership / confirmation popup
                    try:
                        popup_close = page.query_selector(
                            "tp-yt-paper-dialog button[aria-label*='Dismiss' i], ytd-popup-container button[aria-label*='Close' i]"
                        )
                        if popup_close and popup_close.is_visible():
                            popup_close.click()
                    except Exception:
                        pass
                    return True
    except Exception:
        pass
    return False


# --------------------------------------------------------------------------
# Video Candidate Discovery (Trending & Niche)
# --------------------------------------------------------------------------
def fetch_candidate_videos(page, cfg: dict, niche: str, seen: set) -> list[dict]:
    """
    Collects fresh trending and niche videos / Shorts to watch.
    """
    trending_cfg = cfg.get("trending", {})
    trending_enabled = trending_cfg.get("enabled", True)
    content_type = trending_cfg.get("content_type", "both").lower()  # "both", "videos", "shorts"
    cands = []

    # Source 1: YouTube Official Trending Feeds
    if trending_enabled:
        trending_urls = [
            "https://www.youtube.com/feed/trending",
            "https://www.youtube.com/feed/trending?bp=4gINGgt5dG1hX2NoYXJ0cw%3D%3D",  # Music
            "https://www.youtube.com/feed/trending?bp=4gIcGhpnYW1pbmdfY29ycHVzX21vc3RfcG9wdWxhcg%3D%3D",  # Gaming
        ]
        t_url = random.choice(trending_urls)
        print(f"[Feed] Browsing YouTube Trending ({t_url})...")
        try:
            page.goto(t_url, timeout=50000)
            page.wait_for_load_state("domcontentloaded")
            time.sleep(random.uniform(2.5, 4.0))
            _dismiss_banners(page)
            links = page.evaluate("""() => {
                const els = Array.from(document.querySelectorAll('a#video-title, a#video-title-link, a[href*="/shorts/"], ytd-video-renderer a#thumbnail[href*="/watch"]'));
                return els.map(e => ({
                    href: e.href,
                    title: (e.getAttribute('title') || e.textContent || '').trim()
                })).filter(x => x.href && (x.href.includes('/watch') || x.href.includes('/shorts/')));
            }""")
            for l in links:
                href = l["href"].split("&")[0]
                if href and href not in seen:
                    is_short = "/shorts/" in href
                    if content_type == "shorts" and not is_short:
                        continue
                    if content_type == "videos" and is_short:
                        continue
                    cands.append(l)
        except Exception as ex:
            print(f"[Feed] Notice loading trending: {str(ex)[:60]}")

    # Source 2: Niche Search (if configured)
    if niche and not (niche.upper().startswith("PASTE")):
        q = quote(niche)
        search_urls = [
            f"https://www.youtube.com/results?search_query={q}&sp=CAM%253D",  # Sorted by views
            f"https://www.youtube.com/results?search_query={q}",
        ]
        s_url = random.choice(search_urls)
        print(f"[Feed] Browsing Niche Search ({niche})...")
        try:
            page.goto(s_url, timeout=50000)
            page.wait_for_load_state("domcontentloaded")
            time.sleep(random.uniform(2.0, 3.5))
            _dismiss_banners(page)
            links = page.evaluate("""() => {
                const els = Array.from(document.querySelectorAll('a#video-title-link, a#video-title, a[href*="/shorts/"]'));
                return els.map(e => ({
                    href: e.href,
                    title: (e.getAttribute('title') || e.textContent || '').trim()
                })).filter(x => x.href && (x.href.includes('/watch') || x.href.includes('/shorts/')));
            }""")
            for l in links:
                href = l["href"].split("&")[0]
                if href and href not in seen:
                    is_short = "/shorts/" in href
                    if content_type == "shorts" and not is_short:
                        continue
                    if content_type == "videos" and is_short:
                        continue
                    cands.append(l)
        except Exception as ex:
            print(f"[Feed] Notice loading niche search: {str(ex)[:60]}")

    # Deduplicate & shuffle
    unique = []
    for c in cands:
        h = c["href"].split("&")[0]
        if h not in seen:
            seen.add(h)
            unique.append(c)
    random.shuffle(unique)
    print(f"[Feed] Found {len(unique)} candidate videos to watch ({content_type.upper()} mode).")
    return unique


# --------------------------------------------------------------------------
# Excel & Logging
# --------------------------------------------------------------------------
def write_xlsx(rows: list, path: Path, niche: str):
    """Writes styled multi-column Excel worksheet with complete engagement data."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Engagement Worksheet"
    ws.append(["Niche", niche or "Trending", "", "", "", "", "", "Generated", datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append([])
    headers = [
        "#", "Account", "Title", "Channel", "URL", "Views", "Length",
        "Watched (local)", "Suggested action", "Comment Posted",
        "Liked?", "Subscribed?", "Done?",
    ]
    ws.append(headers)

    fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[ws.max_row]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(vertical="center", horizontal="center")

    for i, r in enumerate(rows, 1):
        ws.append([
            i,
            r.get("account", ""),
            r.get("title", ""),
            r.get("channel", ""),
            r.get("url", ""),
            r.get("views", ""),
            fmt_len(r.get("duration")),
            r.get("watched_at", ""),
            r.get("flag", ""),
            r.get("comment", ""),
            "YES" if r.get("liked") else "NO",
            "YES" if r.get("subscribed") else "NO",
            "YES",
        ])

    widths = {
        "A": 5,
        "B": 24,
        "C": 48,
        "D": 26,
        "E": 44,
        "F": 12,
        "G": 10,
        "H": 20,
        "I": 32,
        "J": 42,
        "K": 10,
        "L": 14,
        "M": 8,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"
    wb.save(path)


# --------------------------------------------------------------------------
# Main Engine Runner (callable by CLI or UI thread)
# --------------------------------------------------------------------------
def run_player_engine(cfg: dict = None, status_callback=None, use_real_chrome=None, profile_dir=None, headless=False):
    """
    Core automation loop. Emits state callbacks if provided.
    """
    if cfg is None:
        cfg = load_cfg()

    # use_real_chrome can come from the caller OR from config.json.  Default to the
    # config value when the caller did not explicitly choose (None).
    if use_real_chrome is None:
        use_real_chrome = bool(cfg.get("use_real_chrome", False))

    niche = (cfg.get("niche") or "").strip()
    dedicated = HERE / "browser_profile"

    # Rotation settings
    rot_cfg = cfg.get("account_rotation", {})
    rotation_enabled = rot_cfg.get("enabled", True)
    rot_mins = rot_cfg.get("rotate_minutes", [15, 25])

    # Engagement settings
    eng_cfg = cfg.get("engagement", {})
    auto_like = eng_cfg.get("auto_like", True)
    like_prob = float(eng_cfg.get("like_probability", 0.8))
    auto_comment = eng_cfg.get("auto_comment", True)
    comment_prob = float(eng_cfg.get("comment_probability", 0.5))
    auto_sub = eng_cfg.get("auto_subscribe", True)
    sub_prob = float(eng_cfg.get("subscribe_probability", 0.25))
    comment_pool = eng_cfg.get("comment_pool", DEFAULT_COMMENTS)

    from playwright.sync_api import sync_playwright

    def emit(event_type, **data):
        if status_callback:
            try:
                status_callback({"type": event_type, "timestamp": datetime.now().isoformat(), **data})
            except Exception:
                pass

    print("=" * 68)
    print("Auto-Player: Multi-Account Rotation, Auto-Like, Comment & Subscribe")
    print(f"  Niche              : {niche or '(Global Trending)'}")
    print(f"  Browser Mode       : {'Real Chrome profile' if use_real_chrome else 'Dedicated profile (browser_profile/)'}")
    print(f"  Account Rotation   : {'Enabled (' + str(rot_mins[0]) + ' to ' + str(rot_mins[1]) + ' mins)' if rotation_enabled else 'Disabled (Single account)'}")
    print(f"  Auto-Like          : {'Enabled (prob: ' + str(like_prob) + ')' if auto_like else 'Disabled'}")
    print(f"  Auto-Comment       : {'Enabled (prob: ' + str(comment_prob) + ')' if auto_comment else 'Disabled'}")
    print(f"  Auto-Subscribe     : {'Enabled (prob: ' + str(sub_prob) + ')' if auto_sub else 'Disabled'}")
    print(f"  Watch Duration     : {cfg['play_seconds_per_video']} ± {cfg['play_jitter']} seconds")
    print(f"  Break Frequency    : Every {cfg['break_every']} videos ({cfg['break_seconds'][0]}-{cfg['break_seconds'][1]}s)")
    print(f"  Stop triggers      : Ctrl+C  |  create stop.txt  |  close browser")
    print("=" * 68)

    if STOP_FILE.exists():
        STOP_FILE.unlink()

    outdir = HERE / cfg["output_dir"]
    outdir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    xlsx_path = outdir / ("engagement_%s.xlsx" % stamp)
    log = outdir / "watch_log.csv"
    new_log = not log.exists()
    fh = open(log, "a", encoding="utf-8", newline="")
    if new_log:
        fh.write("time,account,video_url,liked,subscribed,comment,title\n")

    rows = []
    watched_total = 0
    likes_total = 0
    comments_total = 0
    subs_total = 0
    rotations_total = 0

    emit("started", niche=niche, xlsx_path=str(xlsx_path))

    # Browser connection mode
    connect_cdp = bool(cfg.get("connect_cdp", False))
    cdp_url = cfg.get("cdp_url", "http://localhost:9222")

    try:
        # We launch the user's REAL installed Google Chrome (channel='chrome') in every
        # mode here, so Playwright's bundled Chromium is NOT required - only a real Chrome
        # install. No _ensure_browser() gate: it would wrongly block when the user has
        # real Chrome but not Playwright's bundled Chromium.

        with sync_playwright() as p:
            if connect_cdp:
                # ATTACH to a Chrome you are already running (the SAME profile/window you
                # are browsing in). Chrome must be started with --remote-debugging-port=9222;
                # use the provided launch-chrome-debug.bat / .sh helper to start it that way.
                print("Connecting to your running Chrome via CDP:", cdp_url)
                emit("status", status=f"Connecting to your open Chrome at {cdp_url}...")
                try:
                    browser = p.chromium.connect_over_cdp(cdp_url)
                except Exception as e:
                    print("Could not connect to Chrome. Launch it with --remote-debugging-port=9222 (see README).")
                    emit("error", message="Cannot attach to Chrome. Start it with launch-chrome-debug.bat (which launches your Chrome profile on port 9222). Details: %s" % str(e)[:120])
                    return
                # Use the default context that already holds your logged-in session.
                ctx = browser.contexts[0] if browser.contexts else browser.new_context()
                page = ctx.pages[0] if ctx.pages else ctx.new_page()
                emit("status", status="Attached to your open Chrome. Locating YouTube...")
                print("Attached to running Chrome. Navigating to YouTube...")
                page.goto("https://www.youtube.com", wait_until="domcontentloaded", timeout=60000)
                try:
                    _dismiss_banners(page)
                except Exception:
                    pass

            elif use_real_chrome:
                pdir = profile_dir or real_chrome_profile_dir()
                if not pdir:
                    print("Could not detect your Chrome profile. Pass: --profile-dir \"/path/to/Chrome/User Data\"")
                    emit("error", message="Chrome profile not detected")
                    return
                # Which profile INSIDE the User Data folder to use (Default, Profile 5, ...)
                profile_name = cfg.get("chrome_profile_dir", "Default")
                profile_full = os.path.join(pdir, profile_name)
                print("Using real Chrome profile:", pdir, "| profile:", profile_name)
                print("  profile folder:", profile_full)
                emit("status", status="Preparing Chrome profile '%s'..." % profile_name)

                # Direct Playwright attach to the user's LIVE profile is unreliable on
                # Chrome 136 (the profile is owned/locked -> 'Target page, context or browser
                # has been closed' / launch timeouts). Instead, clone the profile's logged-in
                # state into a bot-owned folder and launch THAT. All signed-in accounts are
                # preserved, and the user's real Chrome is never touched.
                print("Copying your Chrome profile '%s' (all signed-in accounts will be preserved)..." % profile_name)
                emit("status", status="Copying your signed-in Chrome profile '%s'..." % profile_name)
                try:
                    launch_dir = clone_real_profile(pdir, profile_name)
                except Exception as e:
                    emit("error", message=(
                        "Could not copy your Chrome profile '%s'. Details: %s"
                        % (profile_name, str(e)[:200])
                    ))
                    return
                print("Cloned profile ready at:", launch_dir)

                print("Launching Chrome on the copied profile...")
                emit("status", status="Opening Chrome with your copied session (accounts preserved)...")
                launch_args = [
                    "--no-first-run",
                    "--no-default-browser-check",
                    "--no-sandbox",
                    "--disable-gpu",
                    "--disable-features=TranslateUI",
                ]
                ctx = None
                last_err = None
                # Retry the launch: the first attempt can time out if a Chrome process was
                # still tearing down; clearing again and retrying usually succeeds.
                for attempt in range(3):
                    try:
                        ctx = p.chromium.launch_persistent_context(
                            str(launch_dir),
                            channel="chrome",
                            headless=False,
                            args=launch_args,
                            viewport=None,
                            locale="en-US",
                            timeout=90000,
                        )
                        break
                    except Exception as e:
                        last_err = e
                        print("  launch attempt %d failed: %s" % (attempt + 1, str(e)[:160]))
                        close_chrome()
                        time.sleep(2)
                if ctx is None:
                    emit("error", message=(
                        "Could not open Chrome on the copied profile. This can happen if the "
                        "profile copy is large. Right-click the bot folder, delete 'cloned_profile', "
                        "then press START BOT again. Details: %s" % str(last_err)[:200]
                    ))
                    return

                # The just-launched window starts on about:blank. Navigate it to YouTube,
                # using 'domcontentloaded' so we don't hang on YouTube's heavy/slow load
                # event. If the first tab is stale, reuse a page that's already pointing at
                # a URL (Chrome may restore tabs), otherwise use the first page.
                page = None
                for cand in (ctx.pages or []):
                    try:
                        url = cand.url or ""
                    except Exception:
                        url = ""
                    if url and not url.startswith("about:"):
                        page = cand
                        break
                if page is None:
                    page = ctx.pages[0] if ctx.pages else ctx.new_page()

                print("Navigating to YouTube...")
                emit("status", status="Loading YouTube in your Chrome profile '%s'..." % profile_name)
                try:
                    page.goto("https://www.youtube.com", wait_until="domcontentloaded", timeout=60000)
                except Exception as e:
                    print("First navigation warning:", str(e)[:120])
                # Let YouTube settle, then dismiss any consent/cookie banners.
                try:
                    page.wait_for_timeout(1500)
                    _dismiss_banners(page)
                except Exception:
                    pass
                emit("status", status="YouTube loaded. Detecting your signed-in accounts...")

            else:
                dedicated.mkdir(exist_ok=True)
                emit("status", status="Opening your saved Chrome session (dedicated profile)...")
                # Launch the user's REAL installed Chrome (channel='chrome') with stealth flags
                # so the saved login is used AND Google's "browser may not be secure" block does
                # not prevent account access. Same profile (browser_profile/) the login window used.
                ctx = launch_persistent_browser(p, dedicated, headless=headless)
                page = ctx.pages[0] if ctx.pages else ctx.new_page()
                page.goto("https://www.youtube.com", wait_until="domcontentloaded", timeout=60000)
                try:
                    _dismiss_banners(page)
                except Exception:
                    pass

            if not use_real_chrome and not headless:
                # Give the user a real window of time to sign in (works in CLI and UI mode).
                wait_for_login(page, status_callback=status_callback)

            # Account roster resolution
            configured_accounts = rot_cfg.get("accounts", [])
            if configured_accounts and len(configured_accounts) > 0:
                account_list = [{"index": idx, "name": name, "handle": ""} for idx, name in enumerate(configured_accounts)]
            else:
                account_list = discover_available_accounts(page)

            if not account_list:
                account_list = [{"index": 0, "name": "Active Account", "handle": ""}]

            emit("accounts_discovered", accounts=account_list)

            current_account_idx = 0
            current_account_name = switch_to_account(page, account_list[0]["index"], account_list[0]["name"])

            # Setup rotation interval (15 to 25 mins)
            min_sec = rot_mins[0] * 60
            max_sec = max(rot_mins[0], rot_mins[1]) * 60
            current_rotation_duration = random.uniform(min_sec, max_sec)
            account_start_time = time.time()

            emit("account_changed", account=current_account_name, index=0, duration=current_rotation_duration)
            print(f"[Timer] Account '{current_account_name}' active. Scheduled rotation in {current_rotation_duration / 60:.1f} minutes.\n")

            seen = set()
            counter = {"n": 0}

            while not stopped():
                if page.is_closed():
                    raise RuntimeError("Browser window was closed by user.")

                # Fetch fresh candidate videos
                candidates = fetch_candidate_videos(page, cfg, niche, seen)
                if not candidates:
                    print("[Feed] No new videos found. Waiting 20 seconds before retry...")
                    emit("status", status="Waiting for new video candidates...")
                    time.sleep(20)
                    continue

                for i, c in enumerate(candidates, 1):
                    if stopped():
                        break
                    if cfg["max_videos"] and counter["n"] >= cfg["max_videos"]:
                        break

                    # Check Account Rotation Timer
                    elapsed = time.time() - account_start_time
                    if rotation_enabled and len(account_list) > 1 and elapsed >= current_rotation_duration:
                        print("\n" + "#" * 65)
                        print(f"[Account Rotation] {elapsed / 60:.1f} minutes completed on account: '{current_account_name}'.")
                        print("[Account Rotation] Rotating to next YouTube account in profile...")
                        print("#" * 65)

                        rotations_total += 1
                        current_account_idx = (current_account_idx + 1) % len(account_list)
                        next_acc = account_list[current_account_idx]

                        emit("account_rotating", current=current_account_name, next=next_acc["name"])
                        time.sleep(random.uniform(5.0, 10.0))
                        current_account_name = switch_to_account(page, next_acc["index"], next_acc["name"])
                        account_start_time = time.time()
                        current_rotation_duration = random.uniform(min_sec, max_sec)

                        emit("account_changed", account=current_account_name, index=current_account_idx, duration=current_rotation_duration)
                        print(f"[Timer] Account '{current_account_name}' active. Scheduled rotation in {current_rotation_duration / 60:.1f} minutes.\n")
                        break

                    per = cfg["play_seconds_per_video"]
                    jit = cfg["play_jitter"]
                    dur = max(15, int(random.uniform(per - jit, per + jit)))
                    video_title = (c.get("title") or c.get("href") or "Video")[:65]

                    print(f"\n▶ [{counter['n'] + 1}] Watching on [{current_account_name}]: {video_title} (~{dur}s)")
                    emit("video_start", title=video_title, url=c["href"], account=current_account_name, duration=dur)

                    try:
                        page.goto(c["href"], timeout=60000)
                        page.wait_for_load_state("domcontentloaded")
                        time.sleep(random.uniform(1.5, 3.5))
                        _dismiss_banners(page)

                        # Click Play if needed
                        try:
                            btn = page.query_selector("button.ytp-large-play-button")
                            if btn and btn.is_visible():
                                btn.click()
                        except Exception:
                            pass

                        # Watch loop with human-like mouse movement and micro-scrolls
                        t0 = time.time()
                        liked_this_video = False
                        comment_posted = ""
                        subscribed_this_channel = False

                        while time.time() - t0 < dur and not stopped():
                            time.sleep(3)
                            try:
                                page.mouse.move(random.randint(200, 1000), random.randint(150, 650), steps=6)
                                if random.random() < 0.25:
                                    page.mouse.wheel(0, random.randint(40, 160))
                            except Exception:
                                break

                            # Engage: Like halfway through video
                            if not liked_this_video and auto_like and (time.time() - t0 >= dur * 0.35):
                                if perform_like(page, like_prob):
                                    liked_this_video = True
                                    likes_total += 1
                                    print(f"     ♥ Liked video via [{current_account_name}]")
                                    emit("action", action="liked", account=current_account_name, title=video_title)

                            # Engage: Comment around 55%
                            if not comment_posted and auto_comment and (time.time() - t0 >= dur * 0.55):
                                comment_posted = perform_comment(page, comment_pool, comment_prob)
                                if comment_posted:
                                    comments_total += 1
                                    print(f"     💬 Commented via [{current_account_name}]: \"{comment_posted[:50]}...\"")
                                    emit("action", action="commented", account=current_account_name, comment=comment_posted, title=video_title)

                            # Engage: Subscribe around 75%
                            if not subscribed_this_channel and auto_sub and (time.time() - t0 >= dur * 0.75):
                                if perform_subscribe(page, sub_prob):
                                    subscribed_this_channel = True
                                    subs_total += 1
                                    print(f"     🔔 Subscribed to creator via [{current_account_name}]")
                                    emit("action", action="subscribed", account=current_account_name, title=video_title)

                        # Collect metadata
                        meta = _scrape_meta(page)
                        row = {
                            "account": current_account_name,
                            "title": (c.get("title") or "")[:200],
                            "url": c["href"],
                            "channel": meta.get("channel", ""),
                            "views": meta.get("views", ""),
                            "duration": meta.get("duration", ""),
                            "watched_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "flag": _flag(meta.get("views")),
                            "liked": liked_this_video,
                            "subscribed": subscribed_this_channel,
                            "comment": comment_posted,
                        }
                        rows.append(row)
                        counter["n"] += 1
                        watched_total = counter["n"]

                        # Save progress to Excel and CSV
                        try:
                            write_xlsx(rows, xlsx_path, niche)
                        except Exception as ex:
                            print(f"   (xlsx update warning: {str(ex)[:60]})")

                        csv_fh_line = '%s,"%s","%s","%s","%s","%s","%s"\n' % (
                            row["watched_at"],
                            row["account"].replace('"', "'"),
                            c["href"].replace('"', "'"),
                            "YES" if row["liked"] else "NO",
                            "YES" if row["subscribed"] else "NO",
                            row["comment"].replace('"', "'"),
                            (c.get("title") or "")[:120].replace('"', "'"),
                        )
                        fh.write(csv_fh_line)
                        fh.flush()

                        print(f"     ✓ Logged | Views: {meta.get('views', '?')} | Length: {fmt_len(meta.get('duration'))} | Liked: {'YES' if liked_this_video else 'NO'} | Subscribed: {'YES' if subscribed_this_channel else 'NO'}")

                        emit("video_complete", video=row, stats={
                            "watched": watched_total,
                            "likes": likes_total,
                            "comments": comments_total,
                            "subs": subs_total,
                            "rotations": rotations_total,
                        })

                        time.sleep(random.uniform(2.5, 5.5))

                        # Scheduled breaks
                        if cfg["break_every"] and counter["n"] % cfg["break_every"] == 0:
                            b = random.uniform(*cfg["break_seconds"])
                            print(f"   ☕ Taking a natural break ({b:.0f}s)...")
                            emit("break", duration=b)
                            time.sleep(b)

                    except Exception as e:
                        print(f"   [Skip] Error during video playback: {str(e)[:80]}")

                if cfg["max_videos"] and counter["n"] >= cfg["max_videos"]:
                    print(f"\nReached max_videos target ({cfg['max_videos']}). Stopping.")
                    break

            try:
                write_xlsx(rows, xlsx_path, niche)
            except Exception:
                pass
            # Only close the context we launched ourselves.  When we attached to your
            # already-running Chrome via CDP we must NOT close it - it's your browser.
            if not connect_cdp:
                try:
                    ctx.close()
                except Exception:
                    pass

    except KeyboardInterrupt:
        print("\nStopped by user (Ctrl+C).")
    except Exception as ex:
        print(f"\nSession finished or stopped: {ex}")
        emit("error", message=str(ex))

    fh.close()
    try:
        write_xlsx(rows, xlsx_path, niche)
    except Exception:
        pass

    emit("finished", stats={
        "watched": watched_total,
        "likes": likes_total,
        "comments": comments_total,
        "subs": subs_total,
        "rotations": rotations_total,
    })

    print("\n" + "=" * 50)
    print(f"Session Finished. Total videos watched: {watched_total}")
    print(f"  Excel Worksheet : {xlsx_path}")
    print(f"  CSV History Log : {log}")
    print("=" * 50)
    if STOP_FILE.exists():
        STOP_FILE.unlink()


def main():
    cfg = load_cfg()

    # CLI Arguments Parsing
    if len(sys.argv) > 2 and "--niche" in sys.argv:
        idx = sys.argv.index("--niche")
        if idx + 1 < len(sys.argv):
            cfg["niche"] = sys.argv[idx + 1].strip()

    use_real_chrome = "--real-chrome" in sys.argv or bool(cfg.get("use_real_chrome", False))
    if "--cdp" in sys.argv:
        cfg["connect_cdp"] = True
        use_real_chrome = True  # attaching to a real Chrome implies real profile
    profile_dir = None
    for i, a in enumerate(sys.argv):
        if a == "--profile-dir" and i + 1 < len(sys.argv):
            profile_dir = sys.argv[i + 1]
        if a == "--profile" and i + 1 < len(sys.argv):
            cfg["chrome_profile_dir"] = sys.argv[i + 1].strip()
        if a == "--cdp-url" and i + 1 < len(sys.argv):
            cfg["cdp_url"] = sys.argv[i + 1]
    headless = "--headless" in sys.argv

    # Override flags
    if "--no-rotate" in sys.argv:
        cfg["account_rotation"]["enabled"] = False
    if "--rotate-min" in sys.argv:
        cfg["account_rotation"]["rotate_minutes"][0] = int(sys.argv[sys.argv.index("--rotate-min") + 1])
    if "--rotate-max" in sys.argv:
        cfg["account_rotation"]["rotate_minutes"][1] = int(sys.argv[sys.argv.index("--rotate-max") + 1])
    if "--no-like" in sys.argv:
        cfg["engagement"]["auto_like"] = False
    if "--no-comment" in sys.argv:
        cfg["engagement"]["auto_comment"] = False
    if "--no-sub" in sys.argv:
        cfg["engagement"]["auto_subscribe"] = False
    if "--shorts-only" in sys.argv:
        cfg["trending"]["content_type"] = "shorts"
    elif "--videos-only" in sys.argv:
        cfg["trending"]["content_type"] = "videos"

    run_player_engine(
        cfg=cfg,
        use_real_chrome=use_real_chrome,
        profile_dir=profile_dir,
        headless=headless,
    )


if __name__ == "__main__":
    main()
