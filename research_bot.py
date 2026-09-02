#!/usr/bin/env python3
"""
Niche Research Bot
==================
Scans OTHER channels in your niche and collects:
  - top channels in the niche (name, handle, subscribers)
  - their recent Shorts (title, url, views, likes, comments, duration, date)
  - fastest-growing videos in the niche
  - top comments on the best-performing videos
  - title/hook keyword analysis

Outputs (in the output/ folder):
  - niche_research_<date>.xlsx   full data, multi-sheet
  - digest_<date>.md             human-readable daily digest

Backends:
  1) Official YouTube Data API v3  -> set "api_key" in config.json (free, recommended)
  2) yt-dlp scraping fallback      -> works with NO key (slower, more rate-limit sensitive)

Run:  python research_bot.py
"""
from __future__ import annotations

import json
import os
import random
import re
import statistics
import sys
import time
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime
from pathlib import Path

HERE = Path(__file__).resolve().parent
CONFIG = HERE / "config.json"
API = "https://www.googleapis.com/youtube/v3"

DEFAULTS = {
    "niche": "",
    "api_key": "",
    "top_channels": 8,
    "videos_per_channel": 15,
    "top_comments_for": 12,
    "output_dir": "output",
    "scrape_max_results": 30,
    "scrape_detail_videos": 12,
}

STOP_WORDS = set(
    "the a an and or of to in for with how why what my me i is are was were this that "
    "you your it its on at by from as be do does not no yes new vs vs."
    .split()
)


# --------------------------------------------------------------------------
# config
# --------------------------------------------------------------------------
def load_cfg() -> dict:
    cfg = dict(DEFAULTS)
    try:
        cfg.update(json.loads(CONFIG.read_text(encoding="utf-8")))
    except FileNotFoundError:
        pass
    if os.environ.get("YOUTUBE_API_KEY"):
        cfg["api_key"] = os.environ["YOUTUBE_API_KEY"]
    return cfg


# --------------------------------------------------------------------------
# helpers
# --------------------------------------------------------------------------
def _get(url: str, key: str, **params):
    params["key"] = key
    req = urllib.request.Request(url + "?" + urllib.parse.urlencode(params))
    with urllib.request.urlopen(req, timeout=30) as r:
        return json.loads(r.read().decode())


def _dur(iso: str) -> int:
    m = re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", iso or "")
    if not m:
        return 0
    h, mi, s = (int(g) if g else 0 for g in m.groups())
    return h * 3600 + mi * 60 + s


def fmt_dur(s) -> str:
    try:
        s = int(s)
    except (TypeError, ValueError):
        return ""
    m, sec = divmod(s, 60)
    h, m = divmod(m, 60)
    return ("%d:%02d:%02d" % (h, m, sec)) if h else ("%d:%02d" % (m, sec))


def video_row(item: dict, channel_title: str) -> dict:
    st = item.get("statistics", {}) or {}
    return {
        "videoId": item["id"],
        "title": (item.get("snippet") or {}).get("title", ""),
        "channel": channel_title,
        "url": "https://www.youtube.com/watch?v=" + item["id"],
        "published": (item.get("snippet") or {}).get("publishedAt", ""),
        "views": int(st.get("viewCount", 0) or 0),
        "likes": int(st.get("likeCount", 0) or 0),
        "comments": int(st.get("commentCount", 0) or 0),
        "duration_s": _dur((item.get("contentDetails") or {}).get("duration", "")),
    }


# --------------------------------------------------------------------------
# backend 1: official YouTube Data API
# --------------------------------------------------------------------------
def api_channels(cfg) -> list:
    key = cfg["api_key"]
    d = _get(API + "/search", key, part="snippet", type="channel",
             q=cfg["niche"], maxResults=cfg["top_channels"])
    items = d.get("items", [])
    if not items:
        return []
    ids = [i["id"]["channelId"] for i in items]
    ch = _get(API + "/channels", key, part="snippet,statistics", id=",".join(ids))
    out = []
    for c in ch.get("items", []):
        s = c.get("snippet", {})
        st = c.get("statistics", {}) or {}
        handle = (s.get("customUrl") or "").lstrip("@")
        out.append({
            "channel": s.get("title", ""),
            "handle": "@" + handle if handle else "",
            "url": ("https://youtube.com/@%s" % handle) if handle
                   else "https://www.youtube.com/channel/" + c["id"],
            "subscribers": int(st.get("subscriberCount", 0) or 0),
            "total_videos": int(st.get("videoCount", 0) or 0),
            "id": c["id"],
        })
    out.sort(key=lambda x: x["subscribers"], reverse=True)
    return out


def api_recent_videos(cfg, channel_id: str, n: int) -> list:
    key = cfg["api_key"]
    d = _get(API + "/search", key, part="snippet", type="video", channelId=channel_id,
             order="date", maxResults=n, videoDuration="short")
    items = d.get("items", [])
    if not items:  # channel may not (only) post shorts
        d = _get(API + "/search", key, part="snippet", type="video", channelId=channel_id,
                 order="date", maxResults=n)
        items = d.get("items", [])
    if not items:
        return []
    ids = [i["id"]["videoId"] for i in items]
    v = _get(API + "/videos", key, part="snippet,statistics,contentDetails", id=",".join(ids))
    return [video_row(it, (it.get("snippet") or {}).get("channelTitle", ""))
            for it in v.get("items", [])]


def api_trending(cfg) -> list:
    key = cfg["api_key"]
    d = _get(API + "/search", key, part="snippet", type="video", q=cfg["niche"],
             videoDuration="short", order="viewCount", maxResults=15)
    items = d.get("items", [])
    if not items:
        return []
    ids = [i["id"]["videoId"] for i in items]
    v = _get(API + "/videos", key, part="snippet,statistics,contentDetails", id=",".join(ids))
    return [video_row(it, (it.get("snippet") or {}).get("channelTitle", ""))
            for it in v.get("items", [])]


def api_comments(cfg, video_id: str) -> list:
    key = cfg["api_key"]
    try:
        d = _get(API + "/commentThreads", key, part="snippet", videoId=video_id,
                 maxResults=5, order="relevance")
    except Exception:
        return []
    out = []
    for t in d.get("items", []):
        sn = t["snippet"]["topLevelComment"]["snippet"]
        out.append({
            "video_id": video_id,
            "author": sn.get("authorDisplayName", ""),
            "text": sn.get("textDisplay", ""),
            "likes": sn.get("likeCount", 0),
            "date": sn.get("publishedAt", ""),
        })
    return out


def run_api(cfg) -> dict:
    channels = api_channels(cfg)
    print("[api] found %d channels" % len(channels))
    videos = []
    for ch in channels[: cfg["top_channels"]]:
        print("[api] recent videos for:", ch["channel"])
        vs = api_recent_videos(cfg, ch["id"], cfg["videos_per_channel"])
        for v in vs:
            v["channel"] = ch["channel"]
        videos.extend(vs)
        time.sleep(1)
    trending = api_trending(cfg)
    print("[api] %d videos total, %d trending" % (len(videos), len(trending)))
    top = sorted(videos + trending, key=lambda v: v["views"], reverse=True)
    top = top[: cfg["top_comments_for"]]
    seen, comments = set(), []
    for v in top:
        if v["videoId"] in seen:
            continue
        seen.add(v["videoId"])
        comments.extend(api_comments(cfg, v["videoId"]))
        time.sleep(0.5)
    return {"channels": channels, "videos": videos,
            "comments": comments, "trending": trending[:15]}


# --------------------------------------------------------------------------
# backend 2: yt-dlp scraping (no key needed)
# --------------------------------------------------------------------------
def run_scrape(cfg) -> dict:
    import yt_dlp

    base = {"quiet": True, "no_warnings": True, "skip_download": True}
    niche = cfg["niche"]
    print("[scrape] searching YouTube for:", niche)
    with yt_dlp.YoutubeDL({**base, "extract_flat": "in_playlist"}) as ydl:
        res = ydl.extract_info(
            "ytsearch%d: %s" % (cfg["scrape_max_results"], niche), download=False)

    rows = []
    for e in res.get("entries", []):
        if not e or not e.get("id"):
            continue
        up = (e.get("upload_date") or "")
        rows.append({
            "videoId": e.get("id"),
            "title": e.get("title", ""),
            "channel": e.get("channel") or e.get("uploader") or "",
            "url": "https://www.youtube.com/watch?v=" + str(e.get("id")),
            "published": up[:8],
            "views": e.get("view_count") or 0,
            "likes": None,
            "comments": None,
            "duration_s": e.get("duration") or 0,
        })

    ch_count = Counter(r["channel"] for r in rows if r["channel"])
    channels = [{"channel": name, "handle": "", "url": "", "subscribers": 0,
                 "total_videos": n, "id": ""}
                for name, n in ch_count.most_common(cfg["top_channels"])]

    ranked = [r for r in rows if r["views"]]
    ranked.sort(key=lambda r: r["views"], reverse=True)
    detail = ranked[: cfg["scrape_detail_videos"]]

    comments = []
    for r in detail:
        print("[scrape] detail:", r["title"][:60])
        try:
            with yt_dlp.YoutubeDL({**base, "getcomments": True,
                                   "extractor_args": {"youtube": {"max_comments": "20"}}}) as ydl2:
                info = ydl2.extract_info(r["url"], download=False)
            up = info.get("upload_date") or ""
            r["likes"] = info.get("like_count")
            r["comments"] = len(info.get("comments") or [])
            r["duration_s"] = info.get("duration") or r["duration_s"]
            if up:
                r["published"] = up[:8]
            for c in (info.get("comments") or [])[:10]:
                comments.append({"video_id": r["videoId"],
                                 "author": c.get("author", ""),
                                 "text": c.get("text", ""),
                                 "likes": c.get("like_count", 0),
                                 "date": ""})
        except Exception as ex:
            print("   ! failed:", str(ex)[:100])
        time.sleep(random.uniform(3, 7))

    return {"channels": channels, "videos": rows,
            "comments": comments, "trending": ranked[:15]}


# --------------------------------------------------------------------------
# analysis
# --------------------------------------------------------------------------
def keyword_analysis(videos: list):
    words = Counter()
    hooks = []
    for v in videos:
        t = v.get("title") or ""
        for w in re.findall(r"[a-z0-9$£₹€]+", t.lower()):
            if len(w) > 2 and w not in STOP_WORDS:
                words[w] += 1
        hooks.append(" ".join(t.split()[:8]))
    return words.most_common(25), [h for h in hooks if h][:20]


# --------------------------------------------------------------------------
# output
# --------------------------------------------------------------------------
def write_excel(path: Path, cfg, data, kw, hooks):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    videos = sorted(data["videos"], key=lambda v: v["views"], reverse=True)
    durations = [v["duration_s"] for v in videos if v.get("duration_s")]
    views = [v["views"] for v in videos if v.get("views")]

    wb = Workbook()
    bold = Font(bold=True)

    def header(ws, row):
        for c in row:
            c.font = bold

    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Niche", cfg["niche"]),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Backend", "YouTube Data API" if data.get("_backend") == "api" else "yt-dlp scrape"),
        ("Channels tracked", len(data["channels"])),
        ("Videos collected", len(videos)),
        ("Top video views", max(views) if views else 0),
        ("Median video views", int(statistics.median(views)) if views else 0),
        ("Median duration", fmt_dur(statistics.median(durations)) if durations else ""),
        ("Top title keywords", ", ".join(w for w, _ in kw[:12])),
    ]
    for r in rows:
        ws.append(r)
    for c in ws[1]:
        c.font = bold
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 90

    ws = wb.create_sheet("Channels")
    ws.append(["Channel", "Handle", "Subscribers", "Total videos on channel", "URL"])
    for c in data["channels"]:
        ws.append([c["channel"], c["handle"], c["subscribers"], c["total_videos"], c["url"]])
    header(ws, ws[1])
    for col, w in zip("ABCDE", (35, 25, 14, 20, 50)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("Videos")
    ws.append(["Title", "Channel", "URL", "Published", "Views", "Likes", "Comments", "Duration"])
    for v in videos:
        pub = v["published"]
        pub = (pub[:10] if isinstance(pub, str) and len(pub) >= 10 else pub or v.get("published", ""))
        ws.append([v["title"], v["channel"], v["url"], pub,
                   v["views"] if v["views"] is not None else "",
                   v["likes"] if v["likes"] is not None else "",
                   v["comments"] if v["comments"] is not None else "",
                   fmt_dur(v["duration_s"])])
    header(ws, ws[1])
    for col, w in zip("ABCD", (60, 30, 48, 12)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("Trending (by views)")
    ws.append(["Title", "Channel", "URL", "Views", "Duration"])
    for v in data["trending"]:
        ws.append([v["title"], v["channel"], v["url"], v["views"], fmt_dur(v["duration_s"])])
    header(ws, ws[1])
    for col, w in zip("ABCD", (60, 30, 48, 12)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("Top Comments")
    ws.append(["Video ID", "Author", "Comment", "Likes"])
    for c in data["comments"]:
        ws.append([c["video_id"], c["author"], c["text"], c["likes"]])
    header(ws, ws[1])
    for col, w in zip("ABCD", (14, 22, 90, 8)):
        ws.column_dimensions[col].width = w

    ws = wb.create_sheet("Keywords & Hooks")
    ws.append(["Keyword", "Count"])
    for w, n in kw:
        ws.append([w, n])
    ws.append([])
    ws.append(["Sample hooks (first 8 words of top titles)"])
    for h in hooks:
        ws.append([h])
    header(ws, ws[1])
    ws.column_dimensions["A"].width = 90
    ws.column_dimensions["B"].width = 10

    wb.save(path)


def write_digest(path: Path, cfg, data, kw, hooks):
    videos = sorted(data["videos"], key=lambda v: v["views"], reverse=True)
    views = [v["views"] for v in videos if v.get("views")]
    durations = [v["duration_s"] for v in videos if v.get("duration_s")]
    L = []
    L.append("# Niche Digest — %s" % cfg["niche"])
    L.append("")
    L.append("Generated: %s" % datetime.now().strftime("%Y-%m-%d %H:%M"))
    L.append("")
    L.append("## Top channels in this niche")
    L.append("")
    for c in data["channels"]:
        subs = (", %.1fK subs" % (c["subscribers"] / 1000.0)) if c.get("subscribers") else ""
        L.append("- **%s** %s — %s" % (c["channel"], subs, c["url"]))
    L.append("")
    L.append("## Top %d videos by views" % min(10, len(videos)))
    L.append("")
    for i, v in enumerate(videos[:10], 1):
        L.append("%d. %s — %s views%s" % (
            i, v["title"], "{:,}".format(v["views"]) if v.get("views") is not None else "?",
            " (channel: %s)" % v["channel"] if v.get("channel") else ""))
        L.append("   %s | length %s" % (v["url"], fmt_dur(v["duration_s"])))
    L.append("")
    L.append("## What's working")
    L.append("")
    if views:
        L.append("- Median views of collected videos: **%s**" % "{:,}".format(int(statistics.median(views))))
    if durations:
        L.append("- Median video length: **%s**" % fmt_dur(statistics.median(durations)))
    if kw:
        L.append("- Most-used title words: **%s**" % ", ".join(w for w, _ in kw[:10]))
    recent = [v for v in videos if str(v.get("published", ""))[:7] ==
              datetime.now().strftime("%Y-%m")]
    if recent:
        L.append("- Videos published this month in sample: **%d**" % len(recent))
    L.append("")
    L.append("## Sample hooks (first 8 words of winning titles)")
    L.append("")
    for h in hooks[:12]:
        L.append("- " + h)
    L.append("")
    L.append("## Top comments (what the audience says)")
    L.append("")
    for c in data["comments"][:15]:
        L.append("- \"%s\" (%s likes — %s)" % (c["text"][:140], c.get("likes", 0), c.get("author", "")))
    L.append("")
    L.append("## Ideas to test (fill in after review)")
    L.append("")
    L.append("- [ ] ")
    L.append("- [ ] ")
    L.append("- [ ] ")
    path.write_text("\n".join(L), encoding="utf-8")


# --------------------------------------------------------------------------
# main
# --------------------------------------------------------------------------
def main():
    cfg = load_cfg()
    if len(sys.argv) > 2 and sys.argv[1] == "--niche":
        cfg["niche"] = sys.argv[2].strip()
    niche = (cfg.get("niche") or "").strip()
    if not niche or "PASTE" in niche.upper():
        print("First set the 'niche' field in config.json (or pass --niche \"your niche\").")
        sys.exit(1)

    outdir = HERE / cfg["output_dir"]
    outdir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")

    data = None
    if cfg.get("api_key"):
        try:
            print("== Using YouTube Data API ==")
            data = run_api(cfg)
            data["_backend"] = "api"
        except Exception as e:
            print("[api] failed (%s) — falling back to scraper" % str(e)[:120])
    if data is None:
        try:
            print("== Using yt-dlp scraper (no API key) ==")
            data = run_scrape(cfg)
            data["_backend"] = "scrape"
        except Exception as e:
            print("Scraper also failed: %s" % e)
            print("Tip: get a free API key from Google Cloud and put it in config.json —")
            print("     it is faster and more reliable. See README.md.")
            sys.exit(1)

    kw, hooks = keyword_analysis(data["videos"] + data["trending"])
    xlsx = outdir / ("niche_research_%s.xlsx" % stamp)
    md = outdir / ("digest_%s.md" % stamp)
    write_excel(xlsx, cfg, data, kw, hooks)
    write_digest(md, cfg, data, kw, hooks)
    print()
    print("DONE")
    print("  Excel : %s" % xlsx)
    print("  Digest: %s" % md)


if __name__ == "__main__":
    main()
